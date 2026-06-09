from functools import lru_cache
import datetime, json, os
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from database import get_db
from models import Asset, Category, Channel, Maintenance, WishItem
from pydantic import BaseModel

router = APIRouter(prefix="/api")

# ------ Pydantic Schemas ------
class AssetCreate(BaseModel):
    name: str
    category_id: Optional[int] = None
    channel_id: Optional[int] = None
    purchase_price: Optional[float] = None
    purchase_date: Optional[str] = None
    current_value: Optional[float] = None
    target_price: Optional[float] = None
    target_date: Optional[str] = None
    status: str = "active"
    currency_code: str = "CNY"
    cover_photo: Optional[str] = None
    image: Optional[str] = None
    notes: Optional[str] = None
    warranty_months: Optional[int] = None
    warranty_start_date: Optional[str] = None
    warranty_end_date: Optional[str] = None
    residual_value: Optional[float] = None

class AssetUpdate(BaseModel):
    name: Optional[str] = None
    category_id: Optional[int] = None
    channel_id: Optional[int] = None
    purchase_price: Optional[float] = None
    purchase_date: Optional[str] = None
    current_value: Optional[float] = None
    target_price: Optional[float] = None
    target_date: Optional[str] = None
    status: Optional[str] = None
    cover_photo: Optional[str] = None
    image: Optional[str] = None
    notes: Optional[str] = None
    warranty_months: Optional[int] = None
    warranty_start_date: Optional[str] = None
    warranty_end_date: Optional[str] = None
    residual_value: Optional[float] = None

class MaintenanceCreate(BaseModel):
    asset_id: int
    date: str
    title: str
    amount: Optional[float] = None
    note: Optional[str] = None

class CategoryCreate(BaseModel):
    name: str
    icon: str = "📦"
    parent_id: Optional[int] = None

class ChannelCreate(BaseModel):
    name: str

class WishCreate(BaseModel):
    name: str
    target_price: Optional[float] = None
    price: Optional[float] = None
    target_date: Optional[str] = None
    note: Optional[str] = None
    category_id: Optional[int] = None
    channel_id: Optional[int] = None

class WishUpdate(BaseModel):
    name: Optional[str] = None
    target_price: Optional[float] = None
    price: Optional[float] = None
    target_date: Optional[str] = None
    note: Optional[str] = None
    category_id: Optional[int] = None
    channel_id: Optional[int] = None
    is_done: Optional[bool] = None

# ------ 品类残值率配置 ------
CATEGORY_DEPRECIATION_RATES = {
    1: 0.15, 7: 0.20, 14: 0.15, 19: 0.15, 23: 0.20,
    29: 0.20, 33: 0.15, 37: 0.10, 42: 0.25, 47: 0.15,
    53: 0.15, 57: 0.20,
}

def get_category_root_id(db, category_id):
    """递归找到顶级分类ID"""
    if not category_id:
        return None
    cat = db.query(Category).filter(Category.id == category_id).first()
    if not cat:
        return None
    while cat.parent_id:
        cat = db.query(Category).filter(Category.id == cat.parent_id).first()
        if not cat:
            return None
    return cat.id

def auto_residual_value(asset, db):
    """根据品类折旧率自动计算残值"""
    if not asset.purchase_price or not asset.purchase_date:
        return None
    days = asset.holding_days
    if days < 365:
        return round(asset.purchase_price * 0.8, 2)  # 第一年内按80%

    root_id = get_category_root_id(db, asset.category_id)
    rate = CATEGORY_DEPRECIATION_RATES.get(root_id, 0.20)
    years = days / 365.0
    # 复利折旧: 价值 = price * (1-rate)^years
    value = asset.purchase_price * ((1 - rate) ** years)
    return round(max(value, 0), 2)

def residual_rate(asset, db):
    """保值率百分比"""
    if not asset.purchase_price or asset.purchase_price == 0:
        return None
    rv = asset.residual_value or auto_residual_value(asset, db) or 0
    rate = (rv / asset.purchase_price) * 100
    return round(rate, 1)

# ------ 辅助函数 ------
def asset_to_dict(a: Asset, db: Session):
    # 预加载
    _ = a.maintenances  # 触发加载
    cat_name = a.category.name if a.category else None
    cat_icon = a.category.icon if a.category else None
    ch_name = a.channel.name if a.channel else None
    return {
        "id": a.id,
        "name": a.name,
        "category_id": a.category_id,
        "category_name": cat_name,
        "category_icon": cat_icon,
        "channel_id": a.channel_id,
        "channel_name": ch_name,
        "purchase_price": a.purchase_price,
        "purchase_date": a.purchase_date.isoformat() if a.purchase_date else None,
        "current_value": a.residual_value if a.residual_value is not None else auto_residual_value(a, db),
        "target_price": a.target_price,
        "target_date": a.target_date.isoformat() if a.target_date else None,
        "status": a.status,
        "currency_code": a.currency_code,
        "cover_photo": a.cover_photo,
        "image": a.image,
        "notes": a.notes,
        "holding_days": a.holding_days,
        "total_maintenance_cost": a.total_maintenance_cost,
        "tco": a.tco,
        "daily_cost": a.daily_cost,
        "net_cost": a.net_cost,
        "warranty_months": a.warranty_months,
        "warranty_start_date": a.warranty_start_date.isoformat() if a.warranty_start_date else None,
        "warranty_end_date": a.warranty_end_date.isoformat() if a.warranty_end_date else None,
        "warranty_status": a.warranty_status,
        "residual_value": a.residual_value,
        "auto_residual_value": auto_residual_value(a, db),
        "residual_rate": residual_rate(a, db),
        "maintenance_count": len(a.maintenances),
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "updated_at": a.updated_at.isoformat() if a.updated_at else None,
    }

# ------ 资产 API ------
@router.get("/assets")
def list_assets(
    search: Optional[str] = Query(None),
    category_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    sort: str = Query("created_at"),
    order: str = Query("desc"),
    db: Session = Depends(get_db)
):
    q = db.query(Asset)
    if search:
        q = q.filter(Asset.name.ilike(f"%{search}%") | (Asset.notes.ilike(f"%{search}%")))
    if category_id:
        q = q.filter(Asset.category_id == category_id)
    if status:
        q = q.filter(Asset.status == status)
    
    sort_map = {
        "created_at": Asset.created_at,
        "name": Asset.name,
        "purchase_date": Asset.purchase_date,
        "purchase_price": Asset.purchase_price,
    }
    sort_col = sort_map.get(sort, Asset.created_at)
    if order == "asc":
        q = q.order_by(sort_col.asc())
    else:
        q = q.order_by(sort_col.desc())
    
    assets = q.all()
    return [asset_to_dict(a, db) for a in assets]

@router.get("/assets/{asset_id}")
def get_asset(asset_id: int, db: Session = Depends(get_db)):
    a = db.query(Asset).options(joinedload(Asset.maintenances)).filter(Asset.id == asset_id).first()
    if not a:
        raise HTTPException(404, "资产不存在")
    result = asset_to_dict(a, db)
    result["maintenances"] = [
        {
            "id": m.id, "date": m.date.isoformat(), "title": m.title,
            "amount": m.amount, "note": m.note
        }
        for m in a.maintenances
    ]
    return result

@router.post("/assets", status_code=201)
def create_asset(data: AssetCreate, db: Session = Depends(get_db)):
    a = Asset(name=data.name)
    for field in ["category_id","channel_id","purchase_price","current_value","target_price","target_date","status","currency_code","cover_photo","image","notes","warranty_months","residual_value"]:
        val = getattr(data, field, None)
        if val is not None:
            setattr(a, field, val)
    if data.purchase_date:
        a.purchase_date = datetime.date.fromisoformat(data.purchase_date)
    if data.warranty_start_date:
        a.warranty_start_date = datetime.date.fromisoformat(data.warranty_start_date)
    db.add(a)
    db.commit()
    db.refresh(a)
    return asset_to_dict(a, db)

@router.put("/assets/{asset_id}")
def update_asset(asset_id: int, data: AssetUpdate, db: Session = Depends(get_db)):
    a = db.query(Asset).filter(Asset.id == asset_id).first()
    if not a:
        raise HTTPException(404, "资产不存在")
    updates = data.model_dump(exclude_unset=True)
    for key, val in updates.items():
        if key == "purchase_date" and val:
            setattr(a, key, datetime.date.fromisoformat(val))
        elif key == "target_date" and val:
            setattr(a, key, datetime.date.fromisoformat(val))
        elif key == "warranty_start_date" and val:
            a.warranty_start_date = datetime.date.fromisoformat(val)
        elif key == "warranty_end_date" and val:
            a.warranty_end_date = datetime.date.fromisoformat(val)
        elif val is not None:
            setattr(a, key, val)
    db.commit()
    db.refresh(a)
    return asset_to_dict(a, db)

@router.delete("/assets/{asset_id}")
def delete_asset(asset_id: int, db: Session = Depends(get_db)):
    a = db.query(Asset).filter(Asset.id == asset_id).first()
    if not a:
        raise HTTPException(404, "资产不存在")
    db.delete(a)
    db.commit()
    return {"ok": True}

# ------ 分类 API ------
@router.get("/categories")
def list_categories(db: Session = Depends(get_db)):
    cats = db.query(Category).filter(Category.parent_id == None).order_by(Category.sort_index).all()
    result = []
    for p in cats:
        child = db.query(Category).filter(Category.parent_id == p.id).order_by(Category.sort_index).all()
        children_data = [{"id": c.id, "name": c.name, "icon": c.icon} for c in child]
        count = db.query(Asset).filter(Asset.category_id == p.id).count() + \
                db.query(Asset).filter(Asset.category_id.in_([c.id for c in child])).count() if child else 0
        result.append({
            "id": p.id, "name": p.name, "icon": p.icon, "children": children_data,
            "asset_count": db.query(Asset).filter(
                Asset.category_id.in_([p.id] + [c.id for c in child])
            ).count()
        })
    return result

@router.post("/categories", status_code=201)
def create_category(data: CategoryCreate, db: Session = Depends(get_db)):
    c = Category(name=data.name, icon=data.icon, parent_id=data.parent_id)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name, "icon": c.icon}

@router.delete("/categories/{cat_id}")
def delete_category(cat_id: int, target_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    c = db.query(Category).filter(Category.id == cat_id).first()
    if not c:
        raise HTTPException(404, "分类不存在")
    # 处理资产迁移
    if target_id:
        db.query(Asset).filter(Asset.category_id == cat_id).update({"category_id": target_id})
        for child in db.query(Category).filter(Category.parent_id == cat_id).all():
            db.query(Asset).filter(Asset.category_id == child.id).update({"category_id": target_id})
    db.delete(c)
    db.commit()
    return {"ok": True}

# ------ 渠道 API ------
@router.get("/channels")
def list_channels(db: Session = Depends(get_db)):
    channels = db.query(Channel).filter(Channel.is_hidden == False).order_by(Channel.sort_index).all()
    return [{"id": c.id, "name": c.name, "is_preset": c.is_preset} for c in channels]

@router.post("/channels", status_code=201)
def create_channel(data: ChannelCreate, db: Session = Depends(get_db)):
    existing = db.query(Channel).filter(Channel.name == data.name).first()
    if existing:
        raise HTTPException(400, "渠道已存在")
    c = Channel(name=data.name)
    db.add(c)
    db.commit()
    db.refresh(c)
    return {"id": c.id, "name": c.name}

# ------ 维护记录 API ------
@router.get("/maintenances/{asset_id}")
def list_maintenances(asset_id: int, db: Session = Depends(get_db)):
    ms = db.query(Maintenance).filter(Maintenance.asset_id == asset_id).order_by(Maintenance.date.desc()).all()
    return [{"id": m.id, "date": m.date.isoformat(), "title": m.title, "amount": m.amount, "note": m.note} for m in ms]

@router.post("/maintenances", status_code=201)
def create_maintenance(data: MaintenanceCreate, db: Session = Depends(get_db)):
    m = Maintenance(asset_id=data.asset_id, title=data.title,
                    date=datetime.date.fromisoformat(data.date))
    if data.amount is not None:
        m.amount = data.amount
    if data.note:
        m.note = data.note
    db.add(m)
    db.commit()
    db.refresh(m)
    return {"id": m.id, "date": m.date.isoformat(), "title": m.title, "amount": m.amount, "note": m.note}

@router.delete("/maintenances/{m_id}")
def delete_maintenance(m_id: int, db: Session = Depends(get_db)):
    m = db.query(Maintenance).filter(Maintenance.id == m_id).first()
    if not m:
        raise HTTPException(404, "维护记录不存在")
    db.delete(m)
    db.commit()
    return {"ok": True}

# ------ 心愿单 API ------
@router.get("/wishes")
def list_wishes(db: Session = Depends(get_db)):
    ws = db.query(WishItem).options(
        joinedload(WishItem.category),
        joinedload(WishItem.channel)
    ).order_by(WishItem.created_at.desc()).all()
    @lru_cache(maxsize=128)
    def _asset_image(asset_id):
        a = db.query(Asset).filter(Asset.id == asset_id).first()
        return a.image if a else None
    return [{"id": w.id, "name": w.name, "target_price": w.target_price,
             "target_date": w.target_date.isoformat() if w.target_date else None,
             "price": w.price,
             "note": w.note, "is_done": w.is_done,
             "category_id": w.category_id,
             "category_name": w.category.name if w.category else None,
             "category_icon": w.category.icon if w.category else None,
             "channel_id": w.channel_id,
             "channel_name": w.channel.name if w.channel else None,
             "converted_asset_id": w.converted_asset_id,
             "image": _asset_image(w.converted_asset_id) if w.converted_asset_id else None,
             "created_at": w.created_at.isoformat() if w.created_at else None} for w in ws]

@router.post("/wishes", status_code=201)
def create_wish(data: WishCreate, db: Session = Depends(get_db)):
    w = WishItem(name=data.name)
    if data.target_price: w.target_price = data.target_price
    if data.price: w.price = data.price
    if data.target_date: w.target_date = datetime.date.fromisoformat(data.target_date)
    if data.note: w.note = data.note
    if data.category_id is not None: w.category_id = data.category_id
    if data.channel_id is not None: w.channel_id = data.channel_id
    db.add(w)
    db.commit()
    db.refresh(w)
    return {"id": w.id, "name": w.name, "target_price": w.target_price,
            "target_date": w.target_date.isoformat() if w.target_date else None,
            "price": w.price,
            "category_id": w.category_id,
            "channel_id": w.channel_id,
            "is_done": False}

@router.post("/wishes/{wish_id}/convert")
def convert_wish_to_asset(wish_id: int, data: AssetCreate, db: Session = Depends(get_db)):
    w = db.query(WishItem).filter(WishItem.id == wish_id).first()
    if not w:
        raise HTTPException(404, "心愿不存在")
    a = Asset(name=data.name)
    for field in ["category_id","channel_id","purchase_price","current_value","notes"]:
        val = getattr(data, field, None)
        if val is not None: setattr(a, field, val)
    # Auto-sync: if user didn't pass category_id/channel_id, inherit from wish
    if data.category_id is None and w.category_id is not None:
        a.category_id = w.category_id
    if data.channel_id is None and w.channel_id is not None:
        a.channel_id = w.channel_id
    if data.purchase_price is None and w.price is not None:
        a.purchase_price = w.price
    if data.current_value is None and w.price is not None:
        a.current_value = w.price
    if data.target_price is None and w.target_price is not None:
        a.target_price = w.target_price
    if data.target_date is None and w.target_date is not None:
        a.target_date = w.target_date
    if data.purchase_date:
        a.purchase_date = datetime.date.fromisoformat(data.purchase_date)
    db.add(a)
    db.flush()
    w.is_done = True
    w.converted_asset_id = a.id
    db.commit()
    db.refresh(a)
    return asset_to_dict(a, db)

@router.put("/wishes/{wish_id}")
def update_wish(wish_id: int, data: WishUpdate, db: Session = Depends(get_db)):
    w = db.query(WishItem).filter(WishItem.id == wish_id).first()
    if not w:
        raise HTTPException(404, "心愿不存在")
    for field in ["name","target_price","price","note","category_id","channel_id","is_done"]:
        val = getattr(data, field, None)
        if val is not None: setattr(w, field, val)
    if data.target_date: w.target_date = datetime.date.fromisoformat(data.target_date)
    db.commit()
    db.refresh(w)
    w = db.query(WishItem).options(
        joinedload(WishItem.category),
        joinedload(WishItem.channel)
    ).filter(WishItem.id == wish_id).first()
    return {"id": w.id, "name": w.name, "target_price": w.target_price,
            "price": w.price,
            "target_date": w.target_date.isoformat() if w.target_date else None,
            "category_id": w.category_id,
            "channel_id": w.channel_id,
            "category_name": w.category.name if w.category else None,
            "channel_name": w.channel.name if w.channel else None,
            "is_done": w.is_done}

@router.delete("/wishes/{wish_id}")
def delete_wish(wish_id: int, db: Session = Depends(get_db)):
    w = db.query(WishItem).filter(WishItem.id == wish_id).first()
    if not w:
        raise HTTPException(404, "心愿不存在")
    db.delete(w)
    db.commit()
    return {"ok": True}


# ------ AI 识别 API ------
import base64
QWEN_API_KEY = os.environ.get("QWEN_API_KEY", "")
QWEN_MODEL = os.environ.get("QWEN_MODEL", "qwen-vl-plus")

@router.post("/ai-recognize")
async def ai_recognize(file: UploadFile = File(...)):
    """接收图片，用千问 VL API 识别产品信息"""
    try:
        contents = await file.read()
        b64 = base64.b64encode(contents).decode()
        mime = file.content_type or "image/jpeg"
        data_url = f"data:{mime};base64,{b64}"

        if QWEN_API_KEY:
            import httpx
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.post(
                    "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {QWEN_API_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": QWEN_MODEL,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "image_url",
                                        "image_url": {"url": data_url}
                                    },
                                    {
                                        "type": "text",
                                        "text": "请识别这张图片中的物品，返回以下 JSON 格式（不要包含 markdown 代码块）：{\"name\": \"物品名称\", \"brand\": \"品牌\", \"category\": \"分类\", \"price\": 预估价格（数字）}"
                                    }
                                ]
                            }
                        ]
                    }
                )
                if resp.status_code == 200:
                    result = resp.json()
                    content = result["choices"][0]["message"]["content"]
                    # 尝试提取 JSON
                    import re
                    json_match = re.search(r'\{[^}]+\}', content, re.DOTALL)
                    if json_match:
                        import json as json_lib
                        parsed = json_lib.loads(json_match.group())
                        return {
                            "name": parsed.get("name", ""),
                            "category": parsed.get("category", ""),
                            "price": parsed.get("price"),
                            "brand": parsed.get("brand", "")
                        }
                    return {"name": "", "category": "", "price": None, "brand": "", "error": "解析响应失败"}
                else:
                    return {"name": "", "category": "", "price": None, "brand": "", "error": f"API 错误: {resp.status_code}"}
        else:
            # 没有配置 API Key，返回空
            return {
                "name": "",
                "category": "",
                "price": None,
                "brand": ""
            }
    except Exception as e:
        return {"name": "", "category": "", "price": None, "brand": "", "error": str(e)}

# ------ Excel 导出 API ------
@router.get("/export/excel")
async def export_excel(db: Session = Depends(get_db)):
    """导出所有资产为 Excel（含图标）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "资产列表"

    # 表头
    headers = ["名称", "图标", "分类", "渠道", "购买价格", "购买日期",
               "当前价值", "状态", "保修月数", "持有天数", "TCO", "日均成本", "备注"]
    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="3A7C6C", end_color="3A7C6C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    assets = db.query(Asset).all()
    for row_idx, a in enumerate(assets, 2):
        ws.cell(row=row_idx, column=1, value=a.name)
        # 图标：用分类的icon
        icon = a.category.icon if a.category else "📦"
        ws.cell(row=row_idx, column=2, value=icon)
        ws.cell(row=row_idx, column=3, value=a.category.name if a.category else "")
        ws.cell(row=row_idx, column=4, value=a.channel.name if a.channel else "")
        ws.cell(row=row_idx, column=5, value=a.purchase_price)
        ws.cell(row=row_idx, column=6, value=a.purchase_date.isoformat() if a.purchase_date else "")
        ws.cell(row=row_idx, column=7, value=a.current_value)
        ws.cell(row=row_idx, column=8, value=a.status)
        ws.cell(row=row_idx, column=9, value=a.warranty_months)
        ws.cell(row=row_idx, column=10, value=a.holding_days)
        ws.cell(row=row_idx, column=11, value=a.tco)
        ws.cell(row=row_idx, column=12, value=a.daily_cost)
        ws.cell(row=row_idx, column=13, value=a.notes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=chiwu_assets_{datetime.date.today().isoformat()}.xlsx"}
    )

# ------ 产品搜索 API ------
PRODUCT_DB_PATH = os.path.join(os.path.dirname(__file__), "product_db.json")

# product_db.json 的分类名称 → 数据库 category.id 映射
PRODUCT_CATEGORY_MAP = {
    "手机": 2, "电脑": 3, "平板": 4,
    "音频": 6, "影音": 6,  # 耳机/音响
    "家电": 7,  # 家电根分类
    "摄影": 15, "相机": 15,  # 相机
    "车辆": 20,  # 汽车
    "家具": 24,  # 沙发/椅
    "游戏": 34,  # 游戏机
    "VR": 34,    # 游戏机 (VR归入游戏娱乐)
    "手表": 38,
    "存储": 50, "外设": 48,
    "充电配件": 51, "网络设备": 52,
}

@router.get("/search-products")
def search_products(q: str = Query("", min_length=0)):
    """本地产品数据库搜索"""
    if not os.path.exists(PRODUCT_DB_PATH):
        return []
    try:
        with open(PRODUCT_DB_PATH, "r", encoding="utf-8") as f:
            products = json.load(f)
    except:
        return []
    
    if not q.strip():
        return products[:20]
    
    q_lower = q.lower().strip()
    results = []
    for p in products:
        score = 0
        name = p.get("name", "").lower()
        brand = p.get("brand", "").lower()
        category = p.get("category", "").lower()
        keywords = [kw.lower() for kw in p.get("keywords", [])]
        
        if q_lower in name:
            score += 10
        if q_lower in brand:
            score += 5
        if q_lower in category:
            score += 3
        for kw in keywords:
            if q_lower in kw:
                score += 2
                break
        
        if score > 0:
            results.append((score, p))
    
    results.sort(key=lambda x: -x[0])
    final = []
    for _, p in results[:15]:
        item = dict(p)
        item["category_id"] = PRODUCT_CATEGORY_MAP.get(item.get("category", ""))
        final.append(item)
    return final

# ------ 统计 API ------
@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    assets = db.query(Asset).all()
    total_assets = len(assets)
    active_assets = sum(1 for a in assets if a.status == "active")
    total_value = sum((a.purchase_price or 0) for a in assets)
    total_maintenance = sum(a.total_maintenance_cost for a in assets)

    # 分类统计
    category_stats = {}
    for a in assets:
        cat_name = a.category.name if a.category else "未分类"
        if cat_name not in category_stats:
            category_stats[cat_name] = {"count": 0, "total_price": 0}
        category_stats[cat_name]["count"] += 1
        category_stats[cat_name]["total_price"] += (a.purchase_price or 0)

    # 日均成本排行（top 10）
    daily_cost_list = []
    for a in assets:
        dc = a.daily_cost
        if dc is not None and dc > 0:
            daily_cost_list.append({"id": a.id, "name": a.name, "daily_cost": dc, "tco": a.tco})

    daily_cost_list.sort(key=lambda x: x["daily_cost"], reverse=True)

    # 月度新增趋势
    monthly = {}
    for a in assets:
        if a.created_at:
            key = a.created_at.strftime("%Y-%m")
            monthly[key] = monthly.get(key, 0) + 1

    return {
        "total_assets": total_assets,
        "active_assets": active_assets,
        "total_value": round(total_value, 2),
        "total_maintenance": round(total_maintenance, 2),
        "category_stats": [{"name": k, **v} for k, v in sorted(category_stats.items(), key=lambda x: -x[1]["total_price"])],
        "daily_cost_top": daily_cost_list[:10],
        "monthly_trend": [{"month": k, "count": v} for k, v in sorted(monthly.items())],
    }
